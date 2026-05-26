import pandas as pd
from sklearn.cluster import KMeans
from sklearn.impute import SimpleImputer
import openpyxl

# 1. Khai bao duong dan file 
duong_dan_file = r"E:\Hoc_tap\Khoa_hoc_du_lieu\BTVN265\diem.xlsx"

# 2. Doc du lieu tu file Excel
du_lieu = pd.read_excel(duong_dan_file, header=None)

# Ham xu ly loi dinh dang ngay thang
def xu_ly_ngay_thang(x):
    if isinstance(x, str) and x.startswith('2026-'):
        parts = x.split('-')
        if len(parts) == 3:
            return float(f"{int(parts[2])}.{int(parts[1])}")
    return x

# 3. Trich xuat thong tin sinh vien va mon hoc
danh_sach_mssv = du_lieu.iloc[1, 3:].values
danh_sach_ten = du_lieu.iloc[2, 3:].values
danh_sach_mon_hoc = du_lieu.iloc[4:, 2].values

# 4. Trich xuat ma tran diem va chuan hoa du lieu
bang_diem = du_lieu.iloc[4:, 3:].map(xu_ly_ngay_thang)
bang_diem = bang_diem.apply(pd.to_numeric, errors='coerce')

# Chuyen vi ma tran
bang_diem_chuyen_vi = bang_diem.T 

# Dien cac diem bi khuyet
bo_dien_khuyet = SimpleImputer(strategy='mean')
bang_diem_da_dien = bo_dien_khuyet.fit_transform(bang_diem_chuyen_vi)

# 5. Thuc hien thuat toan phan cum K-Means
thuat_toan_kmeans = KMeans(n_clusters=3, random_state=42)
nhom_ngau_nhien = thuat_toan_kmeans.fit_predict(bang_diem_da_dien)

# 6. Tinh diem trung binh tich luy de sap xep thu tu nhom
diem_tb_sinh_vien = bang_diem_da_dien.mean(axis=1)

# Tinh diem trung binh cua tung cum
bang_tam = pd.DataFrame({'NhomNgauNhien': nhom_ngau_nhien, 'DiemTB': diem_tb_sinh_vien})
diem_tb_nhom = bang_tam.groupby('NhomNgauNhien')['DiemTB'].mean()

# Sap xep cac cum theo diem trung binh tu cao xuong thap
nhom_da_sap_xep = diem_tb_nhom.sort_values(ascending=False).index

# Anh xa sang nhan so 0, 1, 2 theo thu tu hoc luc
anh_xa_nhom = {
    nhom_da_sap_xep[0]: 0,  
    nhom_da_sap_xep[1]: 1,  
    nhom_da_sap_xep[2]: 2   
}
nhom_chinh_thuc = [anh_xa_nhom[c] for c in nhom_ngau_nhien]

# 7. Tao cau truc bang ket qua 
bang_ket_qua = pd.DataFrame(bang_diem_da_dien, columns=danh_sach_mon_hoc)
bang_ket_qua.insert(0, 'MSSV', danh_sach_mssv)
bang_ket_qua.insert(1, 'Tên Sinh Viên', danh_sach_ten)
bang_ket_qua.insert(2, 'Nhóm Phân Cụm', nhom_chinh_thuc)
bang_ket_qua.insert(3, 'Điểm TB Tích Lũy', diem_tb_sinh_vien.round(2))

# 8. Xuat du lieu ra file Excel va chen ghi chu 
file_dau_ra = "Ket_qua_phan_cum_K58KTP.xlsx"

with pd.ExcelWriter(file_dau_ra, engine='openpyxl') as trinh_ghi_excel:
    bang_ket_qua.to_excel(trinh_ghi_excel, index=False, sheet_name='Ket_qua')
    
    trang_tinh = trinh_ghi_excel.sheets['Ket_qua']
    dong_cuoi = len(bang_ket_qua) + 1
    
    # Ghi chu vao cuoi file Excel bang tieng Viet co dau
    trang_tinh.cell(row=dong_cuoi + 2, column=1, value="Ghi chú:")
    trang_tinh.cell(row=dong_cuoi + 3, column=1, value="0: Giỏi nhất")
    trang_tinh.cell(row=dong_cuoi + 4, column=1, value="1: Khá")
    trang_tinh.cell(row=dong_cuoi + 5, column=1, value="2: Thấp nhất")

print(f"Đã chạy xong! Kết quả phân cụm kèm ghi chú đã được lưu tại: {file_dau_ra}")